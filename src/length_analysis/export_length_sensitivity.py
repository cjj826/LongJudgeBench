# -*- coding: utf-8 -*-
"""
Export length_sensitivity results to xlsx.
Reads outputs/length_sensitivity/*.json, outputs fixed + adaptive bucket summary tables + trend analysis.

Usage:
    python -m src.length_analysis.export_length_sensitivity
"""
import json
import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(BASE_DIR))

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)

FIXED_NAMES = ["<3K", "3-8K", "8-16K", ">16K"]
ADAPTIVE_NAMES = ["Q1", "Q2", "Q3", "Q4"]
DATASETS = ["deepresearch_bench", "realdr", "surge_content", "surge_structure", "verify_bench_hard", "wp_bench", "ma_insights"]
DATASET_LABELS = {
    "deepresearch_bench": "DeepResearch-Bench",
    "realdr": "RealDR",
    "surge_content": "SurGE-Content",
    "surge_structure": "SurGE-Structure",
    "verify_bench_hard": "VerifyBench-Hard",
    "wp_bench": "WP-Bench",
    "ma_insights": "MA-Insights",
}
# Map dataset keys to actual data filenames (some keys differ from filenames)
DATA_FILE_MAP = {
    "surge_content": "surge",
    "surge_structure": "surge",
    "ma_insights": "ma",
}
MODELS = [
    ("qwen3-max", "Qwen3-Max"),
    ("deepseek-v4-flash", "DeepSeek-v4-Flash"),
    ("glm-5.1", "GLM-5.1"),
    ("kimi-k2.6", "Kimi-k2.6"),
]

THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LABEL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
ACC_FMT = Font(name="Consolas", size=11)
TREND_UP_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TREND_DOWN_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")


def add_sheet(wb, title: str, bucket_names: list, suffix: str):
    """Add a bucket result sheet. suffix: "" or "_adaptive" """
    ws = wb.create_sheet(title)

    headers = ["Dataset", "Judge Model"] + bucket_names
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN

    row_idx = 2
    for model_key, model_label in MODELS:
        path = BASE_DIR / "outputs" / "length_sensitivity" / f"{model_key}_vanilla{suffix}.json"
        if not path.exists():
            continue
        with open(path, encoding="utf-8") as f:
            results = json.load(f)

        for ds in DATASETS:
            if ds not in results:
                continue
            cell = ws.cell(row=row_idx, column=1, value=DATASET_LABELS.get(ds, ds))
            cell.fill = LABEL_FILL
            cell.border = THIN

            cell = ws.cell(row=row_idx, column=2, value=model_label)
            cell.border = THIN

            for i, bn in enumerate(bucket_names):
                b = results[ds].get(bn, {})
                acc = b.get("accuracy")
                cell = ws.cell(row=row_idx, column=3 + i)
                if acc is not None:
                    cell.value = round(acc, 4)
                    cell.number_format = "0.0000"
                else:
                    cell.value = "-"
                cell.font = ACC_FMT
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN
            row_idx += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    for idx in range(len(bucket_names)):
        col_letter = chr(ord("C") + idx)
        ws.column_dimensions[col_letter].width = 12


def add_stats_sheet(wb):
    """Length distribution stats sheet."""
    from src.utils.io_utils import load_jsonl
    from src.utils.token_counter import count_tokens

    ws = wb.create_sheet("Length Distribution Stats")

    headers = ["Dataset", "Count", "Min", "Q1", "Median", "Q3", "Max"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN

    for r, ds in enumerate(DATASETS, 2):
        file_key = DATA_FILE_MAP.get(ds, ds)
        path = BASE_DIR / "data_standardized" / f"{file_key}.jsonl"
        if not path.exists():
            continue
        lengths = []
        for rec in load_jsonl(path):
            for resp in rec.get("responses", []):
                lengths.append(count_tokens(resp.get("content", "")))
        if not lengths:
            continue
        lengths.sort()
        n = len(lengths)
        vals = [DATASET_LABELS.get(ds, ds), n, lengths[0],
                lengths[n // 4], lengths[n // 2], lengths[3 * n // 4], lengths[-1]]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c)
            cell.value = val
            cell.border = THIN
            if c >= 3:
                cell.alignment = Alignment(horizontal="right")
            if c == 1:
                cell.fill = LABEL_FILL

    ws.column_dimensions["A"].width = 22
    for col_letter in "BCDEFG":
        ws.column_dimensions[col_letter].width = 12






def _load_all():
    """Load all model JSON data, return {(model_key, fixed|adaptive): {ds: {bn: ...}}}"""
    data = {}
    for model_key, _ in MODELS:
        for suffix, label in [("", "fixed"), ("_adaptive", "adaptive")]:
            path = BASE_DIR / "outputs" / "length_sensitivity" / f"{model_key}_vanilla{suffix}.json"
            if path.exists():
                with open(path, encoding="utf-8") as f:
                    data[(model_key, label)] = json.load(f)
    return data


def _spearman_rank(xs, ys):
    """Compute Spearman rank correlation coefficient (n≥3)."""
    n = len(xs)
    if n < 3:
        return None
    x_ranks = [sorted(xs).index(v) + 1 for v in xs]
    y_ranks = [sorted(ys).index(v) + 1 for v in ys]
    d_sq = sum((xr - yr) ** 2 for xr, yr in zip(x_ranks, y_ranks))
    return 1 - 6 * d_sq / (n * (n * n - 1))


def _kendall_w(rank_matrix):
    """Kendall W coefficient of concordance.
    rank_matrix: m×n, m models across n buckets.
    Returns W (0-1), 1=perfect agreement.
    """
    m = len(rank_matrix)
    n = len(rank_matrix[0])
    if m < 2 or n < 3:
        return None
    # Sum of ranks per bucket
    R = [sum(rank_matrix[j][i] for j in range(m)) for i in range(n)]
    R_mean = sum(R) / n
    S = sum((Ri - R_mean) ** 2 for Ri in R)
    W = 12 * S / (m * m * (n * n * n - n))
    return W


def add_metrics_sheet(wb):
    """Add quantitative metrics sheet."""
    ws = wb.create_sheet("Quantitative Metrics")

    all_data = _load_all()
    available_models = [mk for mk, _ in MODELS if (mk, "adaptive") in all_data]
    bucket_ord = [1, 2, 3, 4]

    # ======= Section 1: Monotonicity & Volatility (Adaptive Buckets) =======
    row_idx = 1
    cell = ws.cell(row=row_idx, column=1, value="A. Monotonicity & Volatility (Adaptive Q1-Q4)")
    cell.font = Font(bold=True, size=12)
    row_idx += 1

    headers = ["Dataset", "Judge Model", "Spearman ρ", "Trend", "ACC Range", "ACC Std", "CV"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN
    row_idx += 1

    for ds in DATASETS:
        first = True
        for mk, ml in MODELS:
            if (mk, "adaptive") not in all_data or ds not in all_data[(mk, "adaptive")]:
                continue
            d = all_data[(mk, "adaptive")][ds]
            vals = [d.get(qn, {}).get("accuracy") for qn in ADAPTIVE_NAMES]
            if all(v is not None for v in vals):
                rho = _spearman_rank(bucket_ord, vals)
                rho_str = f"{rho:.4f}" if rho is not None else "-"
                if rho is not None:
                    trend = "Monotonic Decreasing" if rho < -0.5 else "Monotonic Increasing" if rho > 0.5 else "No Clear Trend"
                else:
                    trend = "-"
                acc_range = max(vals) - min(vals)
                mean_acc = sum(vals) / len(vals)
                acc_std = (sum((v - mean_acc)**2 for v in vals) / len(vals))**0.5
                cv = acc_std / mean_acc if mean_acc > 0 else 0
            else:
                rho_str = "-"; trend = "-"; acc_range = 0; acc_std = 0; cv = 0

            cell = ws.cell(row=row_idx, column=1, value=DATASET_LABELS.get(ds, ds))
            if first: cell.fill = LABEL_FILL
            cell.border = THIN
            ws.cell(row=row_idx, column=2, value=ml).border = THIN
            for ci, v in [(3, rho_str), (4, trend)]:
                ws.cell(row=row_idx, column=ci, value=v).border = THIN
                ws.cell(row=row_idx, column=ci).alignment = Alignment(horizontal="center")
            for ci, v in [(5, acc_range), (6, acc_std), (7, cv)]:
                ws.cell(row=row_idx, column=ci, value=round(v, 4)).border = THIN
                ws.cell(row=row_idx, column=ci).number_format = "0.0000"
                ws.cell(row=row_idx, column=ci).alignment = Alignment(horizontal="center")
            row_idx += 1
            first = False

    # ======= Section 2: Cross-model Spearman Correlation Matrix =======
    row_idx += 1
    cell = ws.cell(row=row_idx, column=1, value="B. Cross-model Spearman Matrix (Adaptive ACC)")
    cell.font = Font(bold=True, size=12)
    row_idx += 1

    for ds in DATASETS:
        model_accs = {}
        for mk in available_models:
            d = all_data.get((mk, "adaptive"), {}).get(ds, {})
            vals = [d.get(qn, {}).get("accuracy") for qn in ADAPTIVE_NAMES]
            if all(v is not None for v in vals):
                model_accs[mk] = vals
        if len(model_accs) < 2:
            continue

        cell = ws.cell(row=row_idx, column=1, value=DATASET_LABELS.get(ds, ds))
        cell.font = Font(bold=True, size=11)
        row_idx += 1

        model_labels = dict(MODELS)
        mk_list = list(model_accs.keys())
        headers = ["Model"] + [model_labels[mk] for mk in mk_list]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=c, value=h)
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center"); cell.border = THIN
        row_idx += 1

        for i, mk_row in enumerate(mk_list):
            ws.cell(row=row_idx, column=1, value=model_labels[mk_row]).fill = LABEL_FILL
            ws.cell(row=row_idx, column=1).border = THIN
            for j, mk_col in enumerate(mk_list):
                if i == j:
                    val = 1.0
                else:
                    rho = _spearman_rank(model_accs[mk_row], model_accs[mk_col])
                    val = rho if rho is not None else 0
                cell = ws.cell(row=row_idx, column=2 + j, value=round(val, 4))
                cell.number_format = "0.0000"; cell.alignment = Alignment(horizontal="center"); cell.border = THIN
                if val > 0.7: cell.fill = TREND_UP_FILL
                elif val < 0.3: cell.fill = TREND_DOWN_FILL
            row_idx += 1
        row_idx += 1

    # ======= Section 3: Kendall W Coefficient of Concordance =======
    cell = ws.cell(row=row_idx, column=1, value="C. Kendall W Coefficient of Concordance")
    cell.font = Font(bold=True, size=12)
    row_idx += 1

    headers = ["Dataset", "Models", "W", "Interpretation"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center"); cell.border = THIN
    row_idx += 1

    for ds in DATASETS:
        rank_matrix = []
        for mk in available_models:
            d = all_data.get((mk, "adaptive"), {}).get(ds, {})
            vals = [d.get(qn, {}).get("accuracy") for qn in ADAPTIVE_NAMES]
            if all(v is not None for v in vals):
                sorted_vals = sorted(vals, reverse=True)
                ranks = [sorted_vals.index(v) + 1 for v in vals]
                rank_matrix.append(ranks)
        if len(rank_matrix) >= 2:
            W = _kendall_w(rank_matrix)
            w_str = f"{W:.4f}" if W is not None else "-"
            expl = "High Agreement" if W is not None and W >= 0.8 else "Moderate Agreement" if W is not None and W >= 0.5 else "Low Agreement" if W is not None else "-"
        else:
            w_str = "-"; expl = "-"

        ws.cell(row=row_idx, column=1, value=DATASET_LABELS.get(ds, ds)).fill = LABEL_FILL
        ws.cell(row=row_idx, column=1).border = THIN
        ws.cell(row=row_idx, column=2, value=len(rank_matrix)).border = THIN
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=w_str).border = THIN
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3).font = Font(bold=True)
        ws.cell(row=row_idx, column=4, value=expl).border = THIN
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")
        row_idx += 1

    ws.column_dimensions["A"].width = 22
    for col_letter in "BCDEFGH":
        ws.column_dimensions[col_letter].width = 16


def main():
    out_dir = BASE_DIR / "outputs" / "length_sensitivity"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "length_sensitivity_summary.xlsx"

    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "Fixed Buckets"
    wb.remove(ws0)

    add_sheet(wb, "Fixed Buckets", FIXED_NAMES, "")
    add_sheet(wb, "Adaptive Buckets", ADAPTIVE_NAMES, "_adaptive")
    add_stats_sheet(wb)
    add_metrics_sheet(wb)

    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"  Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
