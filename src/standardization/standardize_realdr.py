"""
Standardize RealDR (our_data)
Input:
  - data/our_data/task{id}/Doc_*.docx  (640 documents)
  - data/our_data/GT/Human_result1.xlsx  (annotator 1)
  - data/our_data/GT/Human_result2.xlsx  (annotator 2)
  - data/our_data/GT/checklist_mapping_key_T1_T40_1771971296.xlsx  (mapping + weights)
  - data/our_data/task_prompt/task{id}.jsonl  (Instruction)
Output:
  - data_standardized/realdr.jsonl
  - ground_truth/realdr_gt.jsonl
"""

import json
from pathlib import Path

import openpyxl
from docx import Document
from src.standardization.utils import (
    BASE_DIR, ensure_output_dirs, save_jsonl, clear_jsonl, count_tokens
)

# ── Paths ───────────────────────────────────────────────────
OUR_DATA_DIR = BASE_DIR / "data" / "our_data"
GT_DIR = OUR_DATA_DIR / "GT"
MAPPING_XLSX = GT_DIR / "checklist_mapping_key_T1_T40_1771971296.xlsx"
HUMAN1_XLSX = GT_DIR / "Human_result1.xlsx"
HUMAN2_XLSX = GT_DIR / "Human_result2.xlsx"
TASK_PROMPT_DIR = BASE_DIR / "data" / "our_data" / "task_prompt"
BLIND_DOCS_DIR = OUR_DATA_DIR  # task1 ~ task40 all under our_data

OUT_DATA = BASE_DIR / "data_standardized" / "realdr.jsonl"
OUT_GT = BASE_DIR / "ground_truth" / "realdr_gt.jsonl"


def load_mapping():
    """
    Read mapping table.
    Returns: {blind_id: {task_id, model, prompt_pos, weight1, weight2, weight3}}
    """
    wb = openpyxl.load_workbook(MAPPING_XLSX)
    ws = wb["Sheet1"]
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        blind_id = str(row[1]).strip()
        mapping[blind_id] = {
            "task_id": int(row[0]),
            "model": str(row[3]),
            "prompt_pos": int(row[4]),
            "weight1": float(row[6]) if row[6] else 0.33,
            "weight2": float(row[7]) if row[7] else 0.33,
            "weight3": float(row[8]) if row[8] else 0.34,
        }
    return mapping


def load_human_scores(xlsx_path):
    """
    Load scores from a single annotator spreadsheet.
    Returns: {(task_id, blind_id): {dim1: score, dim2: score, dim3: score}}
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    scores = {}
    for row in ws.iter_rows(min_row=3, values_only=True):  # skip header and description rows
        if row[0] is None:
            continue
        blind_id = str(row[2]).strip() if row[2] else ""
        if not blind_id.startswith("Doc_"):   # only process Doc_xxx rows
            continue
        task_id = int(row[1])
        scores[(task_id, blind_id)] = {
            "逻辑结构": float(row[3]) if row[3] else 0,
            "表达形式": float(row[4]) if row[4] else 0,
            "偏见检查": float(row[5]) if row[5] else 0,
        }
    return scores


def load_prompt_cache():
    """Load task_prompt, return {task_id: {pos: context}}"""
    cache = {}
    for task_id in range(1, 41):
        fpath = TASK_PROMPT_DIR / f"task{task_id}.jsonl"
        if not fpath.exists():
            continue
        with open(fpath, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except Exception:
                continue
        cache[task_id] = {}
        for pos in range(1, 5):
            key = f"prompt{pos}"
            if key in data and isinstance(data[key], dict):
                cache[task_id][pos] = data[key].get("context", "")
    return cache


def read_docx_content(filepath):
    """Read docx text content."""
    try:
        doc = Document(str(filepath))
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception:
        return ""


def main():
    ensure_output_dirs()
    clear_jsonl(OUT_DATA)
    clear_jsonl(OUT_GT)

    print("=" * 60)
    print("RealDR Standardization")
    print("=" * 60)

    # Load data
    mapping = load_mapping()
    scores1 = load_human_scores(HUMAN1_XLSX)
    scores2 = load_human_scores(HUMAN2_XLSX)
    prompt_cache = load_prompt_cache()

    print(f"  Mapping entries: {len(mapping)}")
    print(f"  Annotator 1:  {len(scores1)} records")
    print(f"  Annotator 2:  {len(scores2)} records")

    count = 0
    for blind_id, info in mapping.items():
        task_id = info["task_id"]
        model = info["model"]
        prompt_pos = info["prompt_pos"]
        w1, w2, w3 = info["weight1"], info["weight2"], info["weight3"]

        # Instruction
        instruction = prompt_cache.get(task_id, {}).get(prompt_pos, "")

        # Content (docx)
        docx_path = BLIND_DOCS_DIR / f"task{task_id}" / f"{blind_id}.docx"
        content = read_docx_content(docx_path)

        # Ground truth: average of two annotators (fallback to single if annotator 2 is None)
        s1 = scores1.get((task_id, blind_id), {})
        s2 = scores2.get((task_id, blind_id), {})
        has_s2 = any(v is not None and v != 0 for v in s2.values())

        dim_names = ["逻辑结构", "表达形式", "偏见检查"]
        dim_avg = {}
        if has_s2:
            for d in dim_names:
                v1 = s1.get(d, 0) or 0
                v2 = s2.get(d, 0) or 0
                dim_avg[d] = round((v1 + v2) / 2, 1)
        else:
            for d in dim_names:
                dim_avg[d] = s1.get(d, 0) or 0

        annotator_count = 2 if has_s2 else 1

        # Weighted total score
        weighted_total = dim_avg["逻辑结构"] * w1 + dim_avg["表达形式"] * w2 + dim_avg["偏见检查"] * w3
        weighted_total = round(weighted_total, 1)

        # 800-token filter
        if count_tokens(content) < 800:
            continue

        # Standardized data
        record = {
            "dataset": "realdr",
            "id": blind_id,
            "instruction": instruction,
            "prompt_pos": prompt_pos,
            "responses": [{"model": model, "content": content}],
            "ground_truth": {
                "type": "pointwise_multi_dim_weighted",
                "dimensions": dim_avg,
                "weights": {"逻辑结构": w1, "表达形式": w2, "偏见检查": w3},
                "weighted_total": weighted_total,
                "annotators": annotator_count,
            },
            "meta": {
                "language": "zh",
                "task_type": "pointwise",
                "source": "our_data",
                "task_id": task_id,
            },
        }
        save_jsonl(record, OUT_DATA)

        # Ground truth record
        gt_record = {
            "dataset": "realdr",
            "id": blind_id,
            "task_id": task_id,
            "instruction": instruction,
            "prompt_pos": prompt_pos,
            "dimensions": dim_avg,
            "weights": {"逻辑结构": w1, "表达形式": w2, "偏见检查": w3},
            "weighted_total": weighted_total,
            "raw_scores": {
                "annotator1": s1,
            },
            "annotation_type": "pointwise_multi_dim_weighted",
            "annotators": annotator_count,
        }
        if has_s2:
            gt_record["raw_scores"]["annotator2"] = s2
        save_jsonl(gt_record, OUT_GT)
        count += 1

        if count % 100 == 0:
            print(f"  Processed {count}/{len(mapping)}")

    print(f"\n  Total: {count} records")
    print(f"  Output:")
    print(f"    Standardized data: {OUT_DATA}")
    print(f"    Ground truth:      {OUT_GT}")
    print("  Done!")


if __name__ == "__main__":
    main()
