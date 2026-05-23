"""
Reliability results → xlsx export (v2 format)
Generates ACC / Spearman / Kendall / PerQuery_Detail / TokenExceed / ContentFilter sheets.

Column layout (two-level merged headers):
  Model | Framework
    | DR-Bench (Overall | Comprehensiveness | Insight | Instruction Following | Readability)
    | RealDR   (Overall | 逻辑结构 | 表达形式 | 偏见检查)
    | SurGE    (Structure | Content)
    | WP-Bench | Verify
    | MA      (Structure | Insights)
    | Avg.

Usage:
  python scripts/export_results_xlsx.py
  python scripts/export_results_xlsx.py path/to/output.xlsx
"""
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
RESULTS_DIR = BASE_DIR / "outputs" / "reliability_scores"

BASE_PROMPT_MODES = ["vanilla", "vanilla_reference", "vanilla_rubric", "vanilla_rubric_reference"]

# ── Column definitions ──
# (display_name, internal_key, dataset_dir, group_key, pm_prefix, is_pairwise, parent_group)
COLUMNS = [
    # ===== DR-Bench =====
    ("DR-Bench", "dr_bench",      "deepresearch_bench", "all",                   "",                 False, None),
    ("综合性",   "dr_compreh",    "deepresearch_bench", "dim_comprehensiveness",    "",               False, "DR-Bench"),
    ("洞察力",   "dr_insight",    "deepresearch_bench", "dim_insight",              "",               False, "DR-Bench"),
    ("指令遵循", "dr_instfollow", "deepresearch_bench", "dim_instruction_following","",               False, "DR-Bench"),
    ("可读性",   "dr_readable",   "deepresearch_bench", "dim_readability",          "",               False, "DR-Bench"),

    # ===== RealDR =====
    ("RealDR",   "real_dr",       "realdr",          "all",                   "",                 False, None),
    ("逻辑结构", "bd_logic",      "realdr",          "dim_逻辑结构",           "",                 False, "RealDR"),
    ("表达形式", "bd_expr",       "realdr",          "dim_表达形式",           "",                 False, "RealDR"),
    ("偏见检查", "bd_bias",       "realdr",          "dim_偏见检查",           "",                 False, "RealDR"),

    # ===== SurGE =====
    ("Structure","surge_struct",  "surge",              "structure",      "structure_",           False, "SurGE"),
    ("Content",  "surge_content", "surge",              "content",        "content_",             False, "SurGE"),

    # ===== WP-Bench / Verify =====
    ("WP-Bench", "wp_bench",      "wp_bench",           "all",                   "",                 True,  None),
    ("Verify",   "verify",        "verify_bench_hard",  "all",                   "",                 False, None),

    # ===== MA (Insights dimension only) =====
    ("Insights", "ma_insights",  "ma",                "all",            "insights_",            True,  "MA"),
]

# Columns that participate in Avg. (by internal key)
AVG_KEYS = ["dr_bench", "real_dr", "surge_struct", "surge_content",
            "wp_bench", "verify", "ma_insights"]

# Datasets that don't need Spearman/Kendall (pairwise or binary)
NO_RANK_CORR = {"wp_bench", "ma_insights", "verify"}

# Model display order (models not listed default to end)
MODEL_ORDER = [
    "qwen3-32b-nothinking",
    "qwen3-32b",
    "qwen3-max",
    "gpt-4o-mini",
    "gpt-5.2",
    "deepseek-v4-flash",
    "glm-5.1",
    "kimi-k2.6",
]


def find_file(dataset: str, prompt_mode: str, model: str) -> Path:
    return RESULTS_DIR / model / dataset / f"{prompt_mode}.json"


ALL_DATASETS = {"realdr", "deepresearch_bench", "ma", "surge", "verify_bench_hard", "wp_bench"}

def collect_results(filter_models: list = None) -> dict:
    results = {}
    seen_models = set()
    for mdir in RESULTS_DIR.iterdir():
        if not mdir.is_dir():
            continue
        # Skip dataset-level legacy directories (not model dirs)
        if mdir.name in ALL_DATASETS:
            continue
        for fpath in mdir.rglob("*.json"):
            with open(fpath, encoding="utf-8") as f:
                seen_models.add(json.load(f).get("judge_model", "unknown"))

    model_list = sorted(seen_models,
                        key=lambda m: MODEL_ORDER.index(m) if m in MODEL_ORDER else len(MODEL_ORDER))
    # Include models from MODEL_ORDER even without data (e.g. glm-5.1)
    for m in MODEL_ORDER:
        if m not in model_list:
            model_list.insert(MODEL_ORDER.index(m), m)
    if filter_models:
        model_list = [m for m in model_list if m in filter_models]

    for model in model_list:
        for base_pm in BASE_PROMPT_MODES:
            row_key = (model, base_pm)
            results[row_key] = {}
            for _display, ikey, ds, gk, pm_prefix, is_pairwise, _parent in COLUMNS:
                fpath = find_file(ds, pm_prefix + base_pm, model)
                if not fpath.exists():
                    continue
                with open(fpath, encoding="utf-8") as f:
                    data = json.load(f)
                g = data.get("groups", {}).get(gk)
                if g is None:
                    continue
                acc = g.get("accuracy")
                n = g.get("n", 0)
                if n == 0 and acc is None:
                    continue
                pq = data.get("per_query", {}).get(gk, {})
                te = data.get("token_exceed", {})
                cf = data.get("content_filter", {})
                results[row_key][ikey] = {
                    "accuracy": acc,
                    "spearman": g.get("spearman"),
                    "kendall": g.get("kendall"),
                    "n": n,
                    "is_pairwise": is_pairwise,
                    "per_query_spearman": pq.get("spearman"),
                    "per_query_kendall": pq.get("kendall"),
                    "n_queries": pq.get("n_queries_spearman", 0),
                    "token_exceed_error_rate": te.get("error_rate"),
                    "content_filter_error_rate": cf.get("error_rate"),
                }
    return results


def _fmt(val):
    if val is None:
        return "N/A"
    return round(val, 4)


# ── Styles ──
HDR = PatternFill("solid", fgColor="D9E1F2")
SUB = PatternFill("solid", fgColor="E2EFDA")
BORDER = Border(*(Side(style='thin'),)*4)
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)


def write_headers(ws):
    """Write Model|Framework columns + two-level headers + Avg."""
    # Build parent → [col_indices_1based] mapping
    mapping = {}  # parent_name → [col_idx]
    for i, (_display, _ik, _ds, _gk, _pp, _pw, parent) in enumerate(COLUMNS):
        col = i + 3
        if parent is None:
            mapping.setdefault(_display, []).append(col)
        else:
            mapping.setdefault(parent, []).append(col)

    avg_col = len(COLUMNS) + 3
    mapping["Avg."] = [avg_col]

    total_cols = avg_col

    # Row 1 / Row 2: uniform border + fill
    for r in (1, 2):
        for c in range(1, total_cols + 1):
            cl = ws.cell(row=r, column=c)
            cl.border = BORDER
            cl.font = Font(bold=True, size=11)
            cl.alignment = CTR

    # Row 1: parent name
    for pname, cols in mapping.items():
        c1 = cols[0]
        cl = ws.cell(row=1, column=c1, value=pname)
        cl.fill = HDR
        if len(cols) > 1:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=cols[-1])

    # Row 2: child column names
    for i, (_display, _ik, _ds, _gk, _pp, _pw, parent) in enumerate(COLUMNS):
        cl = ws.cell(row=2, column=i+3, value=_display)
        cl.fill = SUB if parent is not None else HDR
    cl = ws.cell(row=2, column=avg_col, value="Avg.")
    cl.fill = HDR

    # Model / Framework freeze
    for r in (1, 2):
        for c, v in [(1, "Model"), (2, "Framework")]:
            cl = ws.cell(row=r, column=c, value=v)
            cl.fill = HDR
            cl.border = BORDER
            cl.font = Font(bold=True, size=11)
            cl.alignment = CTR

    return avg_col


def write_data_rows(ws, results, metric, avg_col):
    """Write data rows. metric: accuracy / spearman / kendall / token_exceed_error_rate"""
    rows = sorted(results, key=lambda r: (
        MODEL_ORDER.index(r[0]) if r[0] in MODEL_ORDER else len(MODEL_ORDER),
        BASE_PROMPT_MODES.index(r[1])))

    for ri, (model, base_pm) in enumerate(rows, start=3):
        ws.cell(row=ri, column=1, value=model).border = BORDER
        ws.cell(row=ri, column=2, value=base_pm).border = BORDER
        rd = results[(model, base_pm)]

        for ci, (_display, ikey, _ds, _gk, _pp, is_pairwise, _parent) in enumerate(COLUMNS):
            col = ci + 3
            cell = ws.cell(row=ri, column=col)
            cell.border = BORDER
            cell.alignment = CTR
            cd = rd.get(ikey)
            if not cd:
                continue
            if metric == "accuracy":
                v = cd.get("accuracy")
                if v is not None:
                    cell.value = _fmt(v)
                    cell.number_format = '0.0000'
            elif metric == "spearman":
                if is_pairwise or ikey in NO_RANK_CORR:
                    cell.value = "N/A"
                    cell.value = "N/A"
                else:
                    v = cd.get("per_query_spearman")
                    if v is not None:
                        cell.value = _fmt(v)
                        cell.number_format = '0.0000'
                    else:
                        # fallback to global spearman (when per-query samples insufficient)
                        v = cd.get("spearman")
                        if v is not None:
                            cell.value = _fmt(v)
                            cell.number_format = '0.0000'
                        else:
                            cell.value = "N/A"
            elif metric == "kendall":
                if is_pairwise or ikey in NO_RANK_CORR:
                    cell.value = "N/A"
                    cell.value = "N/A"
                else:
                    v = cd.get("per_query_kendall")
                    if v is not None:
                        cell.value = _fmt(v)
                        cell.number_format = '0.0000'
                    else:
                        # fallback to global kendall
                        v = cd.get("kendall")
                        if v is not None:
                            cell.value = _fmt(v)
                            cell.number_format = '0.0000'
                        else:
                            cell.value = "N/A"
            else:  # token_exceed_error_rate
                v = cd.get(metric)
                if v is not None:
                    cell.value = _fmt(v)
                    cell.number_format = '0.0000'

        # Avg
        if metric == "accuracy":
            vals = []
            for ak in AVG_KEYS:
                cd = rd.get(ak)
                v = cd.get("accuracy") if cd else None
                if v is not None:
                    vals.append(v)
            if vals:
                cl = ws.cell(row=ri, column=avg_col)
                cl.value = round(sum(vals) / len(vals), 4)
                cl.number_format = '0.0000'
                cl.border = BORDER
                cl.alignment = CTR

        elif metric in ("spearman", "kendall"):
            vals = []
            for ak in AVG_KEYS:
                if ak in NO_RANK_CORR:
                    continue
                cd = rd.get(ak)
                if not cd:
                    continue
                v = cd.get(f"per_query_{metric}")
                if v is not None:
                    vals.append(v)
            if vals:
                cl = ws.cell(row=ri, column=avg_col)
                cl.value = round(sum(vals) / len(vals), 4)
                cl.number_format = '0.0000'
                cl.border = BORDER
                cl.alignment = CTR


def write_perquery(ws, results, avg_col):
    rows = sorted(results, key=lambda r: (
        MODEL_ORDER.index(r[0]) if r[0] in MODEL_ORDER else len(MODEL_ORDER),
        BASE_PROMPT_MODES.index(r[1])))
    for ri, (model, base_pm) in enumerate(rows, start=3):
        ws.cell(row=ri, column=1, value=model).border = BORDER
        ws.cell(row=ri, column=2, value=base_pm).border = BORDER
        rd = results[(model, base_pm)]
        for ci, (_display, ikey, _ds, _gk, _pp, is_pairwise, _parent) in enumerate(COLUMNS):
            col = ci + 3
            cell = ws.cell(row=ri, column=col)
            cell.border = BORDER
            cell.alignment = CTR
            cd = rd.get(ikey)
            if not cd:
                continue
            if is_pairwise or ikey in NO_RANK_CORR:
                cell.value = "N/A"
                continue
            parts = []
            sp = cd.get("per_query_spearman")
            kt = cd.get("per_query_kendall")
            if sp is not None:
                parts.append(f"Sp={_fmt(sp)}")
            if kt is not None:
                parts.append(f"τ={_fmt(kt)}")
            if not parts:
                sp = cd.get("spearman")
                kt = cd.get("kendall")
                if sp is not None:
                    parts.append(f"Sp={_fmt(sp)}")
                if kt is not None:
                    parts.append(f"τ={_fmt(kt)}")
                if parts:
                    parts.append("(global)")
            nq = cd.get("n_queries", 0)
            if nq:
                parts.append(f"q={nq}")
            cell.value = ", ".join(parts) if parts else "N/A"


def export_xlsx(output_path: Path, filter_models: list = None):
    results = collect_results(filter_models=filter_models)
    if not results:
        print("Error: no reliability_scores data found")
        if filter_models:
            print(f"       Specified model(s) {filter_models} may have no data, check paths")
        sys.exit(1)

    wb = Workbook()
    default = wb.active

    SHEETS = [("ACC", "accuracy"), ("Spearman", "spearman"), ("Kendall", "kendall")]
    avg_cols = []
    for name, metric in SHEETS:
        ws = wb.create_sheet(name)
        ac = write_headers(ws)
        write_data_rows(ws, results, metric, ac)
        avg_cols.append(ac)

    ws = wb.create_sheet("PerQuery_Detail")
    ac = write_headers(ws)
    write_perquery(ws, results, ac)

    ws = wb.create_sheet("TokenExceed")
    ac = write_headers(ws)
    write_data_rows(ws, results, "token_exceed_error_rate", ac)

    ws = wb.create_sheet("ContentFilter")
    ac = write_headers(ws)
    write_data_rows(ws, results, "content_filter_error_rate", ac)

    wb.remove(default)

    widths = {1: 20, 2: 26}
    for i in range(3, len(COLUMNS) + 4):
        widths[i] = 12
    for sn in wb.sheetnames:
        ws = wb[sn]
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"已导出: {output_path}")
    print(f"  Sheets: ACC, Spearman, Kendall, PerQuery_Detail, TokenExceed, ContentFilter")
    print(f"  列: Model|Framework|DR-Bench(+4子维)|RealDR(+3子维)|SurGE(Str/Cont)|WP-Bench|Verify|MA(Ins)|Avg.")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Export xlsx summary")
    parser.add_argument("output", nargs="?", default=str(BASE_DIR / "outputs" / "reliability_summary.xlsx"),
                        help="Output xlsx path")
    parser.add_argument("--models", nargs="+", default=None,
                        help="Model names to include (e.g. gpt-4o-mini), defaults to all")
    args = parser.parse_args()
    export_xlsx(Path(args.output), filter_models=args.models)
